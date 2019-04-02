import psycopg2
from openpyxl import load_workbook
import random

# Creating connection engine using Root connection to handle administrative
# functions such as loging in, sending emails in case of missing passords, and
# changing users' passwords.


def root():
    conn = psycopg2.connect("dbname = enterprise user = abdul password = abdul*1983 port = 5432")
    cur = conn.cursor()
    return conn, cur

def connector(usrname, passwd):
    conn = psycopg2.connect("dbname = enterprise user = {} password = {} port = 5432".format(usrname, passwd))
    cur = conn.cursor()
    return conn, cur

def NewUser(newname):
    con, cur = root()
    i = 1
    New_User = newname
    cur.execute('SELECT username FROM users')
    AllUsers = cur.fetchall()
    for oneuser in AllUsers:
        while New_User == oneuser[0]:
            New_User = newname + str(i)
            i += 1
    con.close()
    return New_User

# Logger funtions uses root connection Engine to search for username and password
# If found Logger value returns the boolean value of True, otherwise, it returns
# False.
def Logger(usrname, passwd):
    con, cur = root()
    cur.execute('SELECT username, password FROM users WHERE username = %s and password = %s', (usrname, passwd))
    result = cur.fetchone()
    try:
        if result[0] == usrname and result[1] == passwd:
            return True
    except:
        return False

# .......The following functions are specifically for the user to update his/her own...
# .......information and password.

def GetUser(user):
    con, cur = root()
    cur.execute('SELECT id, firstname, lastname, position, department, email, phone1, phone2, usertype FROM users WHERE username = %s', (user,))
    usrdata = cur.fetchall()
    con.close()
    return usrdata


def UpdateProfile(firstname, lastname, email, phone1, phone2, user):
    con, cur = root()
    cur.execute('UPDATE users SET firstname = %s, lastname = %s, email = %s, phone1 = %s, phone2 = %s WHERE username = %s',(firstname, lastname, email, phone1, phone2, user))
    con.commit()
    con.close()

def ChangePassword(user, currentpswd, newpswd):
    con, cur = root()
    cur.execute('UPDATE users SET password = %s WHERE username = %s AND password = %s', (newpswd, user, currentpswd))
    cur.execute("ALTER ROLE {} WITH PASSWORD '{}'".format(user, newpswd))
    con.commit()
    con.close()

# .......The following functions are for Administrators to create users, get users info,
# .......update users, and activate and deactivate users..
# .......the GetUsers function gets information from all users from database in order to...
# .......display that information in organized tables, and it's different from GetUser function above..

def GetUsers():
    con, cur = root()
    cur.execute('SELECT id, firstname, lastname, username, company, position, department, email, phone1, phone2, usertype, status FROM users ORDER BY id')
    usrdata = cur.fetchall()
    con.close()
    return usrdata

def GetUserInfo(id):
    con, cur = root()
    cur.execute('SELECT * FROM users WHERE id = %s', (id,))
    data = cur.fetchall()
    con.close()
    return data

def UpdateUser(sess_uname, sess_pswd, id, firstname, lastname, company, position, department, email, phone1, phone2, usertype, status):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('UPDATE users SET firstname = %s, lastname = %s, company = %s, position = %s, department = %s, email = %s, phone1 = %s, phone2 = %s, usertype = %s, status = %s WHERE id = %s',
    (firstname, lastname, company, position, department, email, phone1, phone2, usertype, status, id))
    con.commit()
    con.close()

def CreateUser(sess_uname, sess_pswd, firstname, lastname, company, position, department, email, phone1, phone2, usrtype):
    con, cur = connector(sess_uname, sess_pswd)
    #... Adding user to users table...
    UserName = firstname[0].lower() + lastname.lower()
    NewName = NewUser(UserName)
    Password = NewName + '@123'
    cur.execute('INSERT INTO users(firstname, lastname, username, password, company, position, department, email, phone1, phone2, usertype) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
    (firstname, lastname, NewName, Password, company, position, department, email, phone1, phone2, usrtype))
    #... Creating Database User with UserType permissions.
    cur.execute("CREATE ROLE {} WITH LOGIN PASSWORD '{}'".format(NewName, Password))
    con.commit()
    con.close()

def CreateMultipleUsers(sess_uname, sess_pswd, FileName):
    wb = load_workbook(filename = FileName)
    ws = wb['Sheet1']
    i = 2
    while ws.cell(row = i, column = 1).value is not None:
        firstname = str(ws.cell(row = i, column = 1).value)
        lastname = str(ws.cell(row = i, column = 2).value)
        company = str(ws.cell(row = i, column = 3).value)
        position = str(ws.cell(row = i, column = 4).value)
        department = str(ws.cell(row = i, column = 5).value)
        email = str(ws.cell(row = i, column = 6).value)
        phone1 = str(ws.cell(row = i, column = 7).value)
        phone2 = str(ws.cell(row = i, column = 8).value)
        type = str(ws.cell(row = i, column = 9).value)

        CreateUser(sess_uname, sess_pswd, firstname, lastname, company, position, department, email, phone1, phone2, type)
        i += 1

def ResetPassword(sess_uname, sess_pswd, user):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('SELECT username FROM users WHERE id = %s', (user,))
    name = cur.fetchone()
    rst_pass = str(name[0]) + '@123'
    cur.execute('UPDATE users SET password = %s WHERE id = %s' ,(rst_pass, user))
    cur.execute("ALTER ROLE {} WITH PASSWORD '{}'".format(sess_uname, rst_pass))
    con.commit()
    con.close()

def GetProviders():
    con, cur = root()
    cur.execute('SELECT id, name, address, phone_1, phone_2, email, pobox FROM providers')
    data = cur.fetchall()
    con.close()
    return data

def CreateProvider(sess_uname, sess_pswd, name, address, phone1, phone2, email, pobox, description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('INSERT INTO providers(name, address, phone_1, phone_2, email, pobox, description) values(%s, %s, %s, %s, %s, %s, %s)',
    (name, address, phone1, phone2, email, pobox, description))
    con.commit()
    con.close()

def FetchProvider(sess_uname, sess_pswd, prv):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('SELECT * FROM providers WHERE id = %s', (prv,))
    data = cur.fetchall()
    con.close()
    return data

def UpdateProvider(sess_uname, sess_pswd, prv, name, address, phone1, phone2, email, pobox, description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('UPDATE providers SET name = %s, address = %s, phone_1 = %s, phone_2 = %s, email = %s, pobox = %s, description = %s WHERE id = %s',
    (name, address, phone1, phone2, email, pobox, description, prv))
    con.commit()
    con.close()

def ProvidersList():
    con, cur = root()
    cur.execute('SELECT Name FROM Providers')
    provs = cur.fetchall()
    choices = []
    for i in provs:
        choices.append((i[0],i[0]))
    con.close()
    return choices

def GroupList():
    con, cur = root()
    cur.execute('SELECT id, groupname FROM grp ORDER BY id')
    data = cur.fetchall()
    con.close()
    return data

def Groups():
    con, cur = root()
    cur.execute('SELECT groupname FROM grp')
    data = cur.fetchall()
    choices = []
    for i in data:
        choices.append((i[0],i[0]))
    con.close()
    return choices

def GetGroup(id):
    con, cur = root()
    cur.execute('SELECT * FROM grp WHERE id = %s', (id,))
    data = cur.fetchall()
    con.close()
    return data

def AddGroup(sess_uname, sess_pswd, name, description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('INSERT INTO grp(groupname, description) VALUES(%s, %s)', (name, description))
    con.commit()
    con.close()

def UpdateGroup(sess_uname, sess_pswd, id, name, description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('UPDATE grp SET groupname = %s, description = %s WHERE id = %s', (name, description, id))
    con.commit()
    con.close()

def CreateItem(sess_uname, sess_pswd, itemname, brand, provider, unit, uprice, description, size, color, sku, partnum, ieme, length, width, height, diameter, lunit, wunit, hunit, dunit, grp, category):
    code = random.randint(100000000000,999999999999)
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('INSERT INTO Items(Code, Item, Brand, Provider, Unit, Unit_Price, Description, Size, Color, sku, part_number, ieme, lengh, width, height, diameter, l_unit, w_unit, h_unit, d_unit, grp, category) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
    (code, itemname, brand, provider, unit, uprice, description, size, color, sku, partnum, ieme, length, width, height, diameter, lunit, wunit, hunit, dunit, grp, category))
    con.commit()
    con.close()

def UpdateItem(sess_uname, sess_pswd, itm, itemname, brand, provider, unit, uprice, description, size, color, sku, partnum, ieme, length, width, height, diameter, lunit, wunit, hunit, dunit, grp, category):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('UPDATE items SET Item = %s, Brand = %s, Provider = %s, Unit = %s, Unit_Price = %s, Description = %s, Size = %s, Color = %s, sku = %s, part_number = %s, ieme = %s, lengh = %s, width = %s, height = %s, diameter = %s, l_unit = %s, w_unit = %s, h_unit = %s, d_unit = %s, grp = %s, category = %s WHERE id = %s',
    (itemname, brand, provider, unit, uprice, description, size, color, sku, partnum, ieme, length, width, height, diameter, lunit, wunit, hunit, dunit, grp, category, itm))
    con.commit()
    con.close()

def GetItems():
    con, cur = root()
    cur.execute('SELECT * FROM Items ORDER BY ID')
    data = cur.fetchall()
    con.close()
    return data

def FetchItem(itm):
    con, cur = root()
    cur.execute('SELECT * FROM items WHERE id = %s', (itm,))
    data = cur.fetchall()
    con.close()
    return data

def ItemPicker():
    con, cur = root()
    cur.execute('SELECT code, item, provider, unit FROM items')
    data = cur.fetchall()
    con.close()
    return data

def ItemAdder(code):
    con, cur = root()
    cur.execute('SELECT code, item, unit FROM items WHERE code = %s', (code,))
    data = cur.fetchall()
    con.close()
    return data

def GetPackages():
    con, cur = root()
    cur.execute('SELECT packagecode, packagename FROM packages GROUP BY packagecode, packagename ORDER BY packagecode')
    data = cur.fetchall()
    con.close()
    return data

def CreatePackage(sess_uname, sess_pswd, packagename, itemcode, itemname, unit, quantity, description):
    code = 'pkg_' + str(random.randint(100000000000,999999999999))
    con, cur = connector(sess_uname, sess_pswd)
    for i in range(len(itemcode)):
        cur.execute('INSERT INTO packages(packagecode, packagename, itemcode, itemname, unit, quantity, description) VALUES(%s, %s, %s, %s, %s, %s, %s)',
        (code, packagename, itemcode[i], itemname[i], unit[i], quantity[i], description))
        con.commit()

    con.close()

def FetchPackage(pkg):
    con, cur =root()
    #...Get Package Metadata, Pakage code, Name, ad Description.
    cur.execute('SELECT packagecode, packagename, description FROM packages WHERE packagecode = %s GROUP BY packagecode, packagename, description', (pkg,))
    data1 = cur.fetchone()
    #...Get Package content.
    cur.execute('SELECT itemcode, itemname, unit, quantity FROM packages WHERE packagecode = %s', (pkg,))
    data2 = cur.fetchall()
    con.close()
    return data1, data2

def UpdatePackage(sess_uname, sess_pswd, pkg, packagename, itemcode, itemname, unit, quantity, description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('DELETE FROM packages WHERE packagecode = %s', (pkg,))
    for i in range(len(itemcode)):
        cur.execute('INSERT INTO packages(packagecode, packagename, itemcode, itemname, unit, quantity, description) VALUES(%s, %s, %s, %s, %s, %s, %s)',
        (pkg, packagename, itemcode[i], itemname[i], unit[i], quantity[i], description))
        con.commit()
    con.close()

def GetWareHouses():
    con, cur = root()
    cur.execute('SELECT code FROM warehouses ORDER BY code')
    data = cur.fetchall()
    con.close()
    return data

def FetchWarehouse(code):
    con, cur = root()
    cur.execute('SELECT location, code, name, description FROM warehouses WHERE code = %s', (code,))
    data = cur.fetchall()
    con.close()
    return data

def CreateWarehouse(sess_uname, sess_pswd, Name, Code, Location, Description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('INSERT INTO warehouses(location, code, name, description) VALUES(%s, %s, %s, %s)', (Location, Code, Name, Description))
    con.commit()
    con.close()

def UpdateWarehouse(sess_uname, sess_pswd, code, Name, Location, Description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('UPDATE warehouses SET location = %s, name = %s, description = %s WHERE code = %s', (Location, Name, Description, code))
    con.commit()
    con.close()

def CreateBin(sess_uname, sess_pswd, name, code, wh, status, description):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('INSERT INTO bins(code, name, warehouse, description, status) VALUES(%s, %s, %s, %s, %s)', (code, name, wh, description, status))
    con.commit()
    con.close()

def GetBins(code):
    con, cur = root()
    cur.execute('SELECT code, name, status FROM bins WHERE warehouse = %s', (code,))
    data = cur.fetchall()
    con.close()
    return data

def BinInfo(code):
    con, cur = root()
    cur.execute('SELECT code, name, description, status FROM bins WHERE code = %s', (code,))
    data = cur.fetchone()
    con.close()
    return data

def UpdateBin(sess_uname, sess_pswd, code, name, status, desc):
    con, cur = connector(sess_uname, sess_pswd)
    cur.execute('UPDATE bins SET name = %s, description = %s, status = %s WHERE code = %s', (name, desc, status, code))
    con.commit()
    con.close()
#BinInfo('WH_1_1')
